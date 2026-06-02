import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALERT_FILL  = PatternFill("solid", fgColor="FFE0E0")


def _header(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def export_products_excel(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "在庫一覧"
    _header(ws, ["商品コード", "商品名", "カテゴリ", "単位", "在庫数",
                 "アラート在庫数", "原価", "売価", "メモ"])
    for p in products:
        row = [p["code"], p["name"], p["category"], p["unit"], p["stock"],
               p["alert_level"], p["cost_price"], p["sell_price"], p["memo"]]
        ws.append(row)
        if p["alert_level"] > 0 and p["stock"] <= p["alert_level"]:
            for cell in ws[ws.max_row]:
                cell.fill = ALERT_FILL
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or "")) for cell in col
        ) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_logs_excel(logs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入出庫履歴"
    _header(ws, ["日時", "商品コード", "商品名", "区分", "数量",
                 "変動前在庫", "変動後在庫", "理由", "担当者"])
    type_label = {"in": "入庫", "out": "出庫", "adjust": "棚卸調整"}
    for l in logs:
        ws.append([
            l["logged_at"], l["product_code"], l["product_name"],
            type_label.get(l["type"], l["type"]),
            l["quantity"], l["before_stock"], l["after_stock"],
            l["reason"], l["operator"]
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or "")) for cell in col
        ) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
